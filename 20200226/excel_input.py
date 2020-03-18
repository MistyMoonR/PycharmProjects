import ast

import requests
from ast import literal_eval

#链接导入数据
target = 'http://fund.eastmoney.com/data/rankhandler.aspx?op=ph&dt=kf&ft=all&rs=&gs=0&sc=zzf&st=desc&sd=2019-02-26&ed=2020-02-26&qdii=&tabSubtype=,,,,,&pi=1&pn=5208'

req = requests.get(url=target)
data = req.text[22:-160]
data_new = ast.literal_eval(data)
str_list = data_new

import openpyxl as op

wb = op.Workbook()
ws = wb.active


#填表头
col_list = [1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]
row_one_list = ['基金代码', '基金简称', '日期', '单位净值', '累计净值', '日增长率', '近1周', '近1月', '近3月', '近6月', '近1年', '近2年', '近3年', '今年来', '成立来']

for i in range(len(col_list)):
	ws.cell(1, col_list[i]).value = str(row_one_list[i])


row = 2

for count_key in range(len(str_list)):

	for col in range(1, len(str_list[count_key].split(","))):
		ws.cell(row, col).value = (str_list[count_key].split(","))[col - 1]
	row += 1

#从18行开始删除
for i in range(100):
	ws.delete_cols(18)

temp_list = ['周', '月', '三个月', '六个月', '一年', '两年', '三年', '合计']
for col in range(18, 26):
	ws.cell(1, col).value = temp_list[col - 18]

wb.save('总表.xlsx')