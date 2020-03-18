import openpyxl as op

wb = op.load_workbook('近1年.xlsx')
ws = wb.active

for row in range(2, ws.max_row + 1):
	#8到12列
	for col in range(8, 13):
		paiming = ws.max_row - 1

		for temp_row in range(2, ws.max_row + 1):
			if (ws.cell(temp_row, col).value != None) and (ws.cell(row, col).value != None):
				if float(ws.cell(temp_row, col).value) > float(ws.cell(row, col).value):
					paiming = paiming - 1

		ws.cell(row, col + 10).value = paiming

	#算合计分数
	ws.cell(row, 25).value = ws.cell(row, 18).value + ws.cell(row, 19).value +ws.cell(row, 20).value + ws.cell(row, 21).value + ws.cell(row, 22).value


wb.save('近1年排名.xlsx')
