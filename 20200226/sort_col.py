import pandas as pd
import openpyxl as op


def excel_None_rows(x):

	for i in range(1, ws.max_column+1):
		ws.cell(x, i).value = None



def excel_sorted():
	#读取excel
	excel_read = pd.read_excel('近1年排名.xlsx')

	#排序
	excel_sort = excel_read.sort_values(by=['合计'], ascending = False)

	#写入excel
	excel_sort.to_excel(r'近1年合计排序.xlsx', sheet_name = 'sheet1')

	#删除掉排序完成后生成的第一列
	wb = op.load_workbook(filename = '近1年合计排序.xlsx')
	ws = wb.active
	ws.delete_cols(1)

	#删除多余行
	for row in range(52, ws.max_row + 1):
		for i in range(1, ws.max_column+1):
			ws.cell(row, i).value = None

	wb.save('近1年合计排序ff.xlsx')

excel_sorted()

def excel_sorted2():
	#读取excel
	excel_read = pd.read_excel('近2年排名.xlsx')

	#排序
	excel_sort = excel_read.sort_values(by=['合计'], ascending = False)

	#写入excel
	excel_sort.to_excel(r'近2年合计排序.xlsx', sheet_name = 'sheet1')

	#删除掉排序完成后生成的第一列
	wb = op.load_workbook(filename = '近2年合计排序.xlsx')
	ws = wb.active
	ws.delete_cols(1)

	#删除多余行
	for row in range(52, ws.max_row + 1):
		for i in range(1, ws.max_column+1):
			ws.cell(row, i).value = None

	wb.save('近2年合计排序ff.xlsx')

excel_sorted2()


def excel_sorted3():
	#读取excel
	excel_read = pd.read_excel('近3年排名.xlsx')

	#排序
	excel_sort = excel_read.sort_values(by=['合计'], ascending = False)

	#写入excel
	excel_sort.to_excel(r'近3年合计排序.xlsx', sheet_name = 'sheet1')

	#删除掉排序完成后生成的第一列
	wb = op.load_workbook(filename = '近3年合计排序.xlsx')
	ws = wb.active
	ws.delete_cols(1)

	#删除多余行
	for row in range(52, ws.max_row + 1):
		for i in range(1, ws.max_column+1):
			ws.cell(row, i).value = None

	wb.save('近3年合计排序ff.xlsx')

excel_sorted3()
