# -*- coding: utf-8 -*-
import openpyxl as op

wb = op.load_workbook('总表.xlsx')
ws = wb.active


def pick_up_row3(col, str_year):

	row_list = [1]

	for row in range(2, ws.max_row + 1):
		if ws.cell(row, col).value != None:
			row_list.append(row)



	wb1 = op.Workbook()
	ws1 = wb1.active

	ws_row = 1
	for row in row_list:
		for col in range(1, 40):
			ws1.cell(ws_row, col).value = ws.cell(row, col).value
		ws_row = ws_row + 1


	wb1.save('近' + str(str_year) + '年.xlsx')

pick_up_row3(14, 3)


def pick_up_row2(col, str_year):

	row_list = [1]

	for row in range(2, ws.max_row + 1):
		if ws.cell(row, col).value == None:
			if ws.cell(row, col - 1).value != None:
				row_list.append(row)



	wb1 = op.Workbook()
	ws1 = wb1.active

	ws_row = 1
	for row in row_list:
		for col in range(1, 40):
			ws1.cell(ws_row, col).value = ws.cell(row, col).value
		ws_row = ws_row + 1


	wb1.save('近' + str(str_year) + '年.xlsx')

pick_up_row2(14, 2)

def pick_up_row1(col, str_year):

	row_list = [1]

	for row in range(2, ws.max_row + 1):
		if ws.cell(row, col).value == None:
			if ws.cell(row, col - 1).value == None:
				if ws.cell(row, col - 2).value != None:
					row_list.append(row)



	wb1 = op.Workbook()
	ws1 = wb1.active

	ws_row = 1
	for row in row_list:
		for col in range(1, 40):
			ws1.cell(ws_row, col).value = ws.cell(row, col).value
		ws_row = ws_row + 1


	wb1.save('近' + str(str_year) + '年.xlsx')

pick_up_row1(14, 1)





